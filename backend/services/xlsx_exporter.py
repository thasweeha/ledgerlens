"""Audit-linked Excel export (spec §6).

Tab 1: reconciled transaction ledger (schema-shaped per statement type).
Tab 2: statement summary and balance cross-footing metrics.
Every extracted cell carries an OpenPyXL comment pointing back to the visual
evidence its value came from:
  Source: {file} | Page: {n} | Mode: {DIGITAL|RASTER} | BBox: [x0, y0, x1, y1]
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MONEY = "#,##0.00"
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LABEL_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
BALANCED_FILL = PatternFill("solid", fgColor="C6EFCE")
UNBALANCED_FILL = PatternFill("solid", fgColor="FFC7CE")

BANK_HEADERS = [
    "Date",
    "Description",
    "Check Number",
    "Withdrawals/Debits",
    "Deposits/Credits",
    "Balance",
]

CREDIT_CARD_HEADERS = [
    "Trans Date",
    "Post Date",
    "Description",
    "Category/Reference",
    "Amount (Charge/Credit)",
    "Running Balance",
]


def _audit_comment(provenance: dict, file_name: str) -> Optional[Comment]:
    if not provenance:
        return None
    page = provenance.get("page")
    page_num = page + 1 if isinstance(page, int) else 0
    mode = provenance.get("mode") or "UNKNOWN"
    bbox = provenance.get("bbox") or [0, 0, 0, 0]
    coords = ", ".join(f"{float(v):.0f}" for v in bbox)
    text = f"Source: {file_name} | Page: {page_num} | Mode: {mode} | BBox: [{coords}]"
    return Comment(text, "LedgerLens")


def _safe_value(value):
    """Neutralize OCR debris that spreadsheet apps would parse as formulas
    (LibreOffice shows Err:509 for cells like '=' or '=SOMETHING')."""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _money_cell(ws, row: int, col: int, value, provenance=None, file_name: str = ""):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = MONEY
    cell.border = BORDER
    if provenance:
        comment = _audit_comment(provenance, file_name)
        if comment:
            cell.comment = comment
    return cell


def _write_ledger_tab(wb: Workbook, result: Dict) -> None:
    ws = wb.active
    ws.title = "Transaction Ledger"
    file_name = Path(result.get("document", {}).get("source_file", "statement.pdf")).name
    statement_type = result.get("classification", {}).get("statement_type", "BANK_ACCOUNT")

    headers = CREDIT_CARD_HEADERS if statement_type == "CREDIT_CARD" else BANK_HEADERS
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for out_row, txn in enumerate(result.get("transactions", []), start=2):
        prov = txn.get("provenance", {})
        is_cc = statement_type == "CREDIT_CARD"
        amount = txn.get("amount")
        debit = abs(amount) if amount is not None and amount < 0 else None
        credit = amount if amount is not None and amount > 0 else None

        cells: List[tuple] = []
        if is_cc:
            cells = [
                (1, txn.get("date"), prov.get("date")),
                (2, txn.get("post_date"), prov.get("post_date")),
                (3, txn.get("description"), prov.get("description")),
                (4, txn.get("check_number"), prov.get("check_number")),
                (5, amount, prov.get("amount") or prov.get("debit") or prov.get("credit")),
                (6, txn.get("balance"), prov.get("balance")),
            ]
        else:
            cells = [
                (1, txn.get("date"), prov.get("date")),
                (2, txn.get("description"), prov.get("description")),
                (3, txn.get("check_number"), prov.get("check_number")),
                (4, debit, prov.get("debit") or prov.get("amount")),
                (5, credit, prov.get("credit") or prov.get("amount")),
                (6, txn.get("balance"), prov.get("balance")),
            ]

        money_cols = {5, 6} if is_cc else {4, 5, 6}
        for col, value, cell_prov in cells:
            if value is None:
                continue
            if col in money_cols:
                _money_cell(ws, out_row, col, value, cell_prov, file_name)
            else:
                cell = ws.cell(row=out_row, column=col, value=_safe_value(value))
                cell.border = BORDER
                if cell_prov:
                    comment = _audit_comment(cell_prov, file_name)
                    if comment:
                        cell.comment = comment

    widths = [14, 46, 16, 18, 18, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_summary_tab(wb: Workbook, result: Dict) -> None:
    ws = wb.create_sheet("Summary & Reconciliation")
    doc = result.get("document", {})
    account = result.get("account", {})
    balances = result.get("summary_balances", {})
    recon = result.get("reconciliation", {})
    routing = result.get("routing", {})
    classification = result.get("classification", {})

    ws.cell(row=1, column=1, value="STATEMENT SUMMARY & RECONCILIATION").font = TITLE_FONT

    row = 3
    meta = [
        ("Source File", doc.get("source_file")),
        ("Engine Version", doc.get("engine_version")),
        ("Processed At", doc.get("processed_at")),
        ("Statement Type", classification.get("statement_type")),
        ("Classification Confidence", classification.get("confidence")),
        ("Account Number", account.get("account_number")),
        ("Routing Number", account.get("routing_number")),
        ("Statement Period", account.get("statement_period")),
        ("Page Execution Modes", ", ".join(routing.get("modes_per_page", []))),
    ]
    for label, value in meta:
        if value is None:
            continue
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=_safe_value(value))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Balance Cross-Footing").font = LABEL_FONT
    row += 1
    metrics = [
        ("Opening Balance", recon.get("opening_balance"), balances.get("sources", {}).get("opening_balance")),
        ("Closing Balance (reported)", recon.get("closing_balance"), balances.get("sources", {}).get("closing_balance")),
        ("Expected Closing Balance", recon.get("expected_closing_balance"), None),
        ("Difference", recon.get("difference"), None),
    ]
    for key in ("deposits_credits", "withdrawals_debits", "checks_paid",
                "payments_and_credits", "purchases_and_charges", "fees_charged", "interest_charged"):
        if key in recon.get("components", {}):
            metrics.append((key.replace("_", " ").title(), recon["components"][key], None))

    for label, value, source in metrics:
        if value is None:
            continue
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=float(value))
        cell.number_format = MONEY
        if source:
            bbox = source.get("bbox") or [0, 0, 0, 0]
            coords = ", ".join(f"{float(v):.0f}" for v in bbox)
            cell.comment = Comment(
                f"Source: {Path(doc.get('source_file', '')).name} "
                f"| Page: {(source.get('page') or 0) + 1} "
                f"| Mode: {source.get('mode') or 'UNKNOWN'} | BBox: [{coords}]",
                "LedgerLens",
            )
        row += 1

    row += 1
    status = recon.get("status", "UNKNOWN")
    ws.cell(row=row, column=1, value="Verification Status").font = LABEL_FONT
    status_cell = ws.cell(row=row, column=2, value=status)
    status_cell.fill = BALANCED_FILL if status == "BALANCED" else UNBALANCED_FILL
    status_cell.font = Font(bold=True)
    row += 2

    details = [
        ("Running Balance Breaks", recon.get("running_balance_breaks", 0)),
        ("Continuity Checkpoints", len(recon.get("checkpoints", []))),
        ("Ledger Segments", len(recon.get("segments", []))),
        ("Transaction Count", len(result.get("transactions", []))),
        ("Formula", recon.get("formula")),
    ]
    for label, value in details:
        if value is None:
            continue
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=_safe_value(value))
        row += 1

    segments = recon.get("segments") or []
    if len(segments) > 1:
        row += 1
        ws.cell(row=row, column=1, value="Ledger Segments").font = LABEL_FONT
        row += 1
        for seg in segments:
            ws.cell(
                row=row,
                column=1,
                value=f"Segment {seg.get('segment')} (pages {seg.get('opening_page')}-{seg.get('closing_page')})",
            )
            ws.cell(row=row, column=2, value=seg.get("status"))
            ws.cell(row=row, column=3, value=seg.get("difference"))
            ws.cell(row=row, column=4, value=seg.get("transaction_count"))
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


def write_statement_xlsx(result: Dict, path: str | Path) -> Path:
    """Write the unified payload to a two-tab audit workbook."""
    wb = Workbook()
    _write_ledger_tab(wb, result)
    _write_summary_tab(wb, result)
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def payload_from_json(path: str | Path) -> Optional[Dict]:
    """Load a payload dict from the JSON written by `parse --output`."""
    p = Path(path)
    if not p.exists():
        return None
    return json.loads(p.read_text(encoding="utf-8"))
