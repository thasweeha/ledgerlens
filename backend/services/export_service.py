"""
Export Service: Generates structured JSON and multi-tab Excel (.xlsx) reports.
"""
from typing import Dict, Any, Union
import io
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.schemas import StatementPayload


def generate_json_bytes(statement: StatementPayload) -> io.BytesIO:
    """Serializes the verified statement payload into formatted JSON bytes."""
    data = statement.model_dump()
    json_str = json.dumps(data, indent=2, ensure_ascii=False)
    buffer = io.BytesIO(json_str.encode("utf-8"))
    buffer.seek(0)
    return buffer


def generate_xlsx_bytes(statement: StatementPayload) -> io.BytesIO:
    """
    Generates a professionally styled multi-tab Excel workbook:
    - Tab 1: Transactions
    - Tab 2: Reconciliation & Metadata
    """
    wb = openpyxl.Workbook()

    # --- TAB 1: TRANSACTIONS ---
    ws_tx = wb.active
    ws_tx.title = "Transactions"
    ws_tx.views.sheetView[0].showGridLines = True

    # Header styling
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=11)
    credit_font = Font(name="Calibri", size=11, color="166534", bold=True)
    debit_font = Font(name="Calibri", size=11, color="991B1B", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    headers = ["#", "Date", "Description", "Debit", "Credit", "Balance", "Page"]
    ws_tx.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws_tx.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col_idx in (1, 2, 7) else "left")

    for idx, tx in enumerate(statement.transactions, start=1):
        amt_val = float(tx.amount)
        bal_val = float(tx.balance) if tx.balance is not None else ""
        is_credit = tx.type == "credit"
        debit_val = "" if is_credit else abs(amt_val)
        credit_val = abs(amt_val) if is_credit else ""

        row = [
            idx,
            tx.date,
            tx.description,
            debit_val,
            credit_val,
            bal_val,
            tx.page
        ]
        ws_tx.append(row)
        current_row = idx + 1

        # Style data row
        for col_idx in range(1, len(headers) + 1):
            cell = ws_tx.cell(row=current_row, column=col_idx)
            cell.font = cell_font
            cell.border = thin_border

            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (4, 5):
                if cell.value != "":
                    cell.number_format = "$#,##0.00;($#,##0.00);\"-\""
                    cell.alignment = Alignment(horizontal="right")
                    cell.font = credit_font if col_idx == 5 else debit_font
            elif col_idx == 6 and bal_val != "":
                cell.number_format = "$#,##0.00;($#,##0.00);\"-\""
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal="center")

    # --- TAB 2: RECONCILIATION SUMMARY ---
    ws_summary = wb.create_sheet(title="Reconciliation Summary")
    ws_summary.views.sheetView[0].showGridLines = True

    rec = statement.reconciliation

    ws_summary.append(["LedgerLens Statement Reconciliation Audit Report"])
    ws_summary.merge_cells("A1:C1")
    title_cell = ws_summary.cell(row=1, column=1)
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1E293B")
    ws_summary.append([])  # blank row

    summary_headers = ["Audit Metric", "Value", "Status"]
    ws_summary.append(summary_headers)
    for col_idx in range(1, 4):
        cell = ws_summary.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    status_pass = rec.reconciled
    status_label = "RECONCILED (PASS)" if status_pass else "OUT OF BALANCE (FAIL)"
    status_fill = PatternFill(
        start_color="DCFCE7" if status_pass else "FEE2E2",
        end_color="DCFCE7" if status_pass else "FEE2E2",
        fill_type="solid"
    )
    status_font = Font(
        name="Calibri", size=11, bold=True,
        color="166534" if status_pass else "991B1B"
    )

    metrics = [
        ("Filename", statement.filename, ""),
        ("Statement Pages", statement.page_count, ""),
        ("Total Transactions Extracted", rec.transaction_count, ""),
        ("Opening Balance", f"${rec.opening_balance:,.2f}", ""),
        ("Total Credits (+)", f"${rec.total_credits:,.2f}", ""),
        ("Total Debits (-)", f"${rec.total_debits:,.2f}", ""),
        ("Calculated Closing Balance", f"${rec.calculated_closing:,.2f}", ""),
        ("Stated Closing Balance", f"${rec.closing_balance:,.2f}", ""),
        ("Difference", f"${rec.difference:,.2f}", ""),
        ("Ledger Reconciliation Status", status_label, "PASS" if status_pass else "FAIL")
    ]

    for m_label, m_val, m_stat in metrics:
        ws_summary.append([m_label, m_val, m_stat])
        row_idx = ws_summary.max_row
        for col_idx in range(1, 4):
            c = ws_summary.cell(row=row_idx, column=col_idx)
            c.font = cell_font
            c.border = thin_border

        if m_label == "Ledger Reconciliation Status":
            c_val = ws_summary.cell(row=row_idx, column=2)
            c_val.fill = status_fill
            c_val.font = status_font

    # Auto-fit column widths across both sheets
    for ws in (ws_tx, ws_summary):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or "")
                max_len = max(max_len, len(val_str))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
