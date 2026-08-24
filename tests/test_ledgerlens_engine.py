"""
Test suite for LedgerLens (hybrid pipeline, backend API, exports, and CLI surface).
"""
import io
import os
import json
from decimal import Decimal
from pathlib import Path

import pytest
from PIL import Image, ImageDraw
import pymupdf
from fastapi.testclient import TestClient

from pipeline.money import parse_amount_decimal
from pipeline.router import route_document
from backend.app import app


client = TestClient(app)


def create_sample_pdf(is_scanned: bool = False) -> bytes:
    """Generates an in-memory PDF statement for testing."""
    doc = pymupdf.open()
    page = doc.new_page(width=595, height=842)  # A4 points

    if not is_scanned:
        # Insert digital vector text
        text = (
            "FIRST GLOBAL BANK - STATEMENT\n"
            "Account: 123456789\n"
            "Opening Balance: $5,000.00\n"
            "\n"
            "Date Description Amount Balance\n"
            "2024-01-05 DEPOSIT CLIENT PAYMENT $2,500.00 $7,500.00\n"
            "2024-01-12 OFFICE SUPPLIES PURCHASE -$150.00 $7,350.00\n"
            "2024-01-20 UTILITIES ELECTRICITY -$350.00 $7,000.00\n"
            "2024-01-28 CONSULTING FEE REVENUE $1,000.00 $8,000.00\n"
            "\n"
            "Closing Balance: $8,000.00\n"
        )
        page.insert_text((50, 70), text, fontsize=12)
    else:
        # Insert a raster image to simulate scanned document
        img = Image.new("RGB", (1200, 1600), color="white")
        draw = ImageDraw.Draw(img)
        draw.text((50, 50), "FIRST GLOBAL BANK - SCANNED", fill="black")
        draw.text((50, 100), "Opening Balance: $1,000.00", fill="black")
        draw.text((50, 200), "2024-02-01 INVOICE PAYMENT $500.00", fill="black")
        draw.text((50, 300), "2024-02-15 SOFTWARE LICENSE -$100.00", fill="black")
        draw.text((50, 400), "Closing Balance: $1,400.00", fill="black")

        img_bytes = io.BytesIO()
        img.save(img_bytes, format="PNG")
        rect = pymupdf.Rect(0, 0, 595, 842)
        page.insert_image(rect, stream=img_bytes.getvalue())

    pdf_bytes = doc.tobytes()
    doc.close()
    return pdf_bytes


def test_money_parsing():
    """Tests the pipeline's negative-notation and currency handling."""
    assert parse_amount_decimal("$1,250.50") == Decimal("1250.50")
    assert parse_amount_decimal("-$45.00") == Decimal("-45.00")
    assert parse_amount_decimal("($100.25)") == Decimal("-100.25")
    assert parse_amount_decimal("45.00 CR") == Decimal("45.00")
    assert parse_amount_decimal("45.00 DR") == Decimal("-45.00")
    assert parse_amount_decimal("+$12.00") == Decimal("12.00")
    assert parse_amount_decimal("not a number") is None


def test_router_routes_digital_and_scanned(tmp_path):
    """Tests page-mode routing for vector and raster statements."""
    vec_path = tmp_path / "vector.pdf"
    vec_path.write_bytes(create_sample_pdf(is_scanned=False))
    plan = route_document(vec_path)
    assert plan["page_count"] == 1
    assert plan["modes"] == ["DIGITAL_VECTOR"]

    scan_path = tmp_path / "scanned.pdf"
    scan_path.write_bytes(create_sample_pdf(is_scanned=True))
    plan = route_document(scan_path)
    assert plan["page_count"] == 1
    assert plan["modes"] == ["FLATBED_RASTER"]


def test_engine_digital_end_to_end(tmp_path):
    """Runs the unified engine over a digital vector statement (no OCR model needed)."""
    from pipeline.engine import process_statement

    pdf_path = tmp_path / "statement.pdf"
    pdf_path.write_bytes(create_sample_pdf(is_scanned=False))

    payload = process_statement(pdf_path)
    assert payload["document"]["page_count"] == 1
    assert payload["routing"]["digital_pages"] == [0]
    assert isinstance(payload["transactions"], list)
    assert payload["classification"]["statement_type"]
    assert payload["reconciliation"]["status"] in {
        "BALANCED", "UNBALANCED", "INSUFFICIENT_DATA"
    }
    for txn in payload["transactions"]:
        assert set(txn) >= {"index", "page", "date", "description", "amount", "type"}


def test_export_service_json_and_xlsx(tmp_path):
    """Tests JSON export and multi-tab Excel workbook generation."""
    from backend.schemas import (
        StatementPayload, TransactionRow, ReconciliationSummary
    )
    from backend.services.export_service import generate_json_bytes, generate_xlsx_bytes

    stmt = StatementPayload(
        session_id="test-session",
        filename="statement.pdf",
        page_count=1,
        pages=[],
        opening_balance=1000.0,
        closing_balance=1500.0,
        transactions=[
            TransactionRow(date="2024-01-01", description="Client Pay",
                           amount=500.0, type="credit", page=1)
        ],
        reconciliation=ReconciliationSummary(
            reconciled=True,
            opening_balance=1000.0,
            closing_balance=1500.0,
            total_credits=500.0,
            total_debits=0.0,
            calculated_closing=1500.0,
            difference=0.0,
            transaction_count=1
        )
    )

    json_buf = generate_json_bytes(stmt)
    data = json.loads(json_buf.read().decode("utf-8"))
    assert data["opening_balance"] == 1000.0
    assert data["reconciliation"]["reconciled"] is True

    xlsx_buf = generate_xlsx_bytes(stmt)
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_buf.read()))
    assert "Transactions" in wb.sheetnames
    assert "Reconciliation Summary" in wb.sheetnames


def test_fastapi_endpoints():
    """Tests all FastAPI backend endpoints (re-OCR skipped unless TrOCR enabled)."""
    # 1. Health check
    res_health = client.get("/api/health")
    assert res_health.status_code == 200
    assert res_health.json()["status"] == "healthy"

    # 2. Upload vector PDF
    vec_pdf = create_sample_pdf(is_scanned=False)
    files = {"file": ("statement.pdf", vec_pdf, "application/pdf")}
    res_upload = client.post("/api/upload", files=files)
    assert res_upload.status_code == 200
    payload = res_upload.json()
    session_id = payload["session_id"]
    assert session_id is not None
    assert payload["page_count"] == 1
    assert payload["pages"][0]["type"] in ("vector", "scanned")

    # 3. Retrieve page image
    res_img = client.get(f"/api/page-image/{session_id}/0")
    assert res_img.status_code == 200
    assert res_img.headers["content-type"] == "image/jpeg"
    assert len(res_img.content) > 1000

    # 4. Session retrieval
    res_sess = client.get(f"/api/session/{session_id}")
    assert res_sess.status_code == 200
    assert res_sess.json()["session_id"] == session_id

    # 5. Reconcile calculation endpoint
    res_recon = client.post("/api/reconcile", json={
        "opening_balance": 1000.0,
        "closing_balance": 1750.0,
        "transactions": [
            {"id": "t1", "date": "2024-01-01", "description": "Deposit", "amount": 1000.0, "type": "credit", "page": 1},
            {"id": "t2", "date": "2024-01-02", "description": "Purchase", "amount": 250.0, "type": "debit", "page": 1}
        ]
    })
    assert res_recon.status_code == 200
    recon_data = res_recon.json()
    assert recon_data["reconciled"] is True
    assert recon_data["calculated_closing"] == 1750.0

    # 6. Export JSON
    res_exp_json = client.post("/api/export", json={
        "statement": payload,
        "format": "json"
    })
    assert res_exp_json.status_code == 200
    assert "application/json" in res_exp_json.headers["content-type"]

    # 7. Export XLSX
    res_exp_xlsx = client.post("/api/export", json={
        "statement": payload,
        "format": "xlsx"
    })
    assert res_exp_xlsx.status_code == 200
    assert "openxmlformats" in res_exp_xlsx.headers["content-type"]

    # 8. Serve UI Root index.html
    res_root = client.get("/")
    assert res_root.status_code == 200
    assert "LEDGERLENS" in res_root.text.upper()


@pytest.mark.skipif(
    os.environ.get("LEDGERLENS_TEST_TROCR") != "1",
    reason="Downloads TrOCR model weights; enable with LEDGERLENS_TEST_TROCR=1"
)
def test_reocr_endpoint_with_trocr():
    """Exercises /api/re-ocr with the real TrOCR recognizer."""
    vec_pdf = create_sample_pdf(is_scanned=False)
    files = {"file": ("statement.pdf", vec_pdf, "application/pdf")}
    res_upload = client.post("/api/upload", files=files)
    session_id = res_upload.json()["session_id"]

    res_ocr = client.post("/api/re-ocr", json={
        "session_id": session_id,
        "page_index": 0,
        "bbox": {"x": 50, "y": 70, "width": 2000, "height": 400, "page": 1},
        "target_field": "all"
    })
    assert res_ocr.status_code == 200
    body = res_ocr.json()
    assert set(body) >= {"raw_text", "cleaned_text", "parsed_date", "parsed_amount", "target_field", "confidence"}


if __name__ == "__main__":
    pytest.main(["-v", __file__])
